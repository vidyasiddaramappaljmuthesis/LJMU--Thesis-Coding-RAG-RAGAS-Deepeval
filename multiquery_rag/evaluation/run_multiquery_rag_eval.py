"""
End-to-end Multi-Query RAG evaluation: RAGAS + DeepEval -> Excel

All metrics are REFERENCE-BASED (zero LLM judge calls during evaluation).
Golden dataset provides expected_answer + expected_source_ids as ground truth.
Each query makes exactly 2 Groq calls:
  Step 1 : query expansion  (1 Groq call)
  Step 3 : answer generation (1 Groq call)

Pipeline per query:
  Step 1  : Groq query expansion     (NUM_QUERY_VARIANTS variants, 1 Groq call)
  Step 2  : Multi-retrieval          (PER_QUERY_TOP_K docs per variant, no LLM)
  Step 3  : RRF fusion               (merge + deduplicate to FINAL_TOP_K, no LLM)
  Step 4  : Groq answer generation   (1 Groq call — second LLM call)
  Step 5  : RAGAS metrics            (reference-based, no LLM)
  Step 6  : DeepEval metrics         (reference-based, no LLM)

RAGAS metrics  : Faithfulness, AnswerRelevancy, ContextPrecision,
                 ContextRecall, FactualCorrectness
DeepEval metrics: AnswerRelevancy, Faithfulness, ContextualPrecision,
                  ContextualRecall, ContextualRelevancy, Hallucination

Metric computation (no LLM judge):
  Faithfulness        : sentence-level support (fraction of answer sentences
                        whose meaningful tokens are >= 50% covered by context)
  AnswerRelevancy     : TF-IDF cosine(generated_answer, question)
  ContextPrecision    : Average-Precision@k  (doc ID in expected_source_ids)
  ContextRecall       : token-recall(expected_answer tokens in contexts)
  FactualCorrectness  : ROUGE-L F1(generated_answer, expected_answer)
  DE ContextualRelevancy : mean TF-IDF cosine(each retrieved doc, question)
  DE Hallucination    : 1.0 - sentence_faithfulness

Note: 2 Groq calls per query -> use longer inter-query delays (10-25 s).

Excel output (5 sheets):
  1. Combined Results  -- every query as a row, every metric as a column
  2. RAG Responses     -- retrieved contexts per query (fused top-k)
  3. RAGAS Metrics     -- per-query RAGAS scores
  4. DeepEval Metrics  -- per-query DeepEval scores + computation method
  5. Summary           -- mean / min / max / std for all metrics

Usage:
    conda run --no-capture-output -n rag_eval python multiquery_rag/evaluation/run_multiquery_rag_eval.py
    conda run --no-capture-output -n rag_eval python multiquery_rag/evaluation/run_multiquery_rag_eval.py --limit 5
    conda run --no-capture-output -n rag_eval python multiquery_rag/evaluation/run_multiquery_rag_eval.py --limit 10 --batch-size 2
"""

import sys, json, re, argparse, logging, textwrap, threading, random, time
from pathlib import Path
from typing import Any, List
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity as _cosine_sim

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from dotenv import load_dotenv
load_dotenv(PROJECT_ROOT / ".env")

from multiquery_rag.implementation.ingestion import build_vector_store, get_collection
from multiquery_rag.implementation.retriever import retrieve_for_query
from multiquery_rag.implementation.fusion import rrf_fuse
from multiquery_rag.implementation.config import (
    GROQ_API_KEYS, GROQ_MODEL, NUM_QUERY_VARIANTS, PER_QUERY_TOP_K, FINAL_TOP_K,
    RRF_K, EXPANDER_TEMPERATURE,
)

GOLDEN_PATH = PROJECT_ROOT / "dataset" / "golden" / "golden_dataset.csv"
RAG_NAME    = "MultiQuery-RAG"


def _default_output() -> Path:
    from datetime import datetime
    ts = datetime.now().strftime("%d-%m-%Y_%I-%M%p")
    return PROJECT_ROOT / "multiquery_rag" / "results" / f"{RAG_NAME}_{ts}.xlsx"


logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s | %(levelname)s | %(message)s",
                    datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

_W = 90  # output width

RAGAS_COLS = [
    "faithfulness",
    "answer_relevancy",
    "context_precision",
    "context_recall",
    "factual_correctness(mode=f1)",
]
DE_COLS = [
    "AnswerRelevancyMetric",
    "FaithfulnessMetric",
    "ContextualPrecisionMetric",
    "ContextualRecallMetric",
    "ContextualRelevancyMetric",
    "HallucinationMetric",
]
RAGAS_DISPLAY = {
    "faithfulness":                  "RAGAS_Faithfulness",
    "answer_relevancy":              "RAGAS_AnswerRelevancy",
    "context_precision":             "RAGAS_ContextPrecision",
    "context_recall":                "RAGAS_ContextRecall",
    "factual_correctness(mode=f1)":  "RAGAS_FactualCorrectness",
}
DE_DISPLAY = {
    "AnswerRelevancyMetric":     "DE_AnswerRelevancy",
    "FaithfulnessMetric":        "DE_Faithfulness",
    "ContextualPrecisionMetric": "DE_ContextualPrecision",
    "ContextualRecallMetric":    "DE_ContextualRecall",
    "ContextualRelevancyMetric": "DE_ContextualRelevancy",
    "HallucinationMetric":       "DE_Hallucination",
}
DE_REASONS = {
    "AnswerRelevancyMetric":     "TF-IDF cosine(generated_answer, question)",
    "FaithfulnessMetric":        "sentence-support(ans sentences in retrieved_contexts)",
    "ContextualPrecisionMetric": "AP@k (ID match vs expected_source_ids)",
    "ContextualRecallMetric":    "token-recall(expected_answer tokens in retrieved_contexts)",
    "ContextualRelevancyMetric": "mean TF-IDF cosine(each_retrieved_doc, question)",
    "HallucinationMetric":       "1.0 - sentence_faithfulness",
}
RAGAS_METHODS = {
    "faithfulness":                  "sentence-support(ans sentences in ctx)",
    "answer_relevancy":              "TF-IDF cosine(ans, question)",
    "context_precision":             "AP@k (ID match vs expected_source_ids)",
    "context_recall":                "token-recall(exp_ans tokens in ctx)",
    "factual_correctness(mode=f1)":  "ROUGE-L F1(ans, exp_ans)",
}


# ══════════════════════════════════════════════════════════════════════════════
# Parallel key routing
# ══════════════════════════════════════════════════════════════════════════════

_QUERIES_PER_KEY   = 20
_PRIMARY_KEY_COUNT = 5

_exhausted_keys: set = set()
_exhausted_lock      = threading.Lock()
_print_lock          = threading.Lock()


def _primary_key_idx(zero_based_pos: int, batch_size: int = None) -> int:
    sz = batch_size if batch_size else _QUERIES_PER_KEY
    return min(zero_based_pos // sz, _PRIMARY_KEY_COUNT - 1)


def _ordered_keys(primary_idx: int) -> List[str]:
    result: List[str] = []
    if primary_idx < len(GROQ_API_KEYS):
        k = GROQ_API_KEYS[primary_idx]
        with _exhausted_lock:
            if k not in _exhausted_keys:
                result.append(k)
    for i in range(_PRIMARY_KEY_COUNT, len(GROQ_API_KEYS)):
        k = GROQ_API_KEYS[i]
        with _exhausted_lock:
            if k not in _exhausted_keys:
                result.append(k)
    return result


def _groq_chat(primary_idx: int, messages: list, temperature: float = 0.0,
               max_tokens: int = 1024) -> str:
    """Key-managed Groq call used for both expansion and generation."""
    import groq as _g
    attempts = _ordered_keys(primary_idx)
    if not attempts:
        raise RuntimeError("No available Groq keys left")
    for key in attempts:
        try:
            kw = dict(model=GROQ_MODEL, messages=messages,
                      temperature=temperature, max_tokens=max_tokens)
            r = _g.Groq(api_key=key).chat.completions.create(**kw)
            return r.choices[0].message.content
        except Exception as e:
            err       = str(e)
            err_lower = err.lower()
            if "organization_restricted" in err_lower:
                with _exhausted_lock:
                    _exhausted_keys.add(key)
                with _print_lock:
                    log.warning("Key ...%s org-restricted -> permanently banned", key[-6:])
            elif "tokens per day" in err_lower or ("tpd" in err_lower and "limit" in err_lower):
                with _exhausted_lock:
                    _exhausted_keys.add(key)
                with _print_lock:
                    log.warning("Key ...%s daily TPD exhausted -> banned for session", key[-6:])
            elif "400" in err and "organization" in err_lower:
                with _exhausted_lock:
                    _exhausted_keys.add(key)
                with _print_lock:
                    log.warning("Key ...%s 400/org error -> permanently banned", key[-6:])
            else:
                with _print_lock:
                    log.warning("Key ...%s transient limit -> skipping", key[-6:])
    raise RuntimeError("All available Groq keys exhausted")


_EXPANDER_SYSTEM = (
    "You are a query rewriting assistant. "
    "Given a user question, produce exactly {n} different ways to ask the same question. "
    "Each variant must preserve the original intent but use different vocabulary or structure. "
    "Output ONLY a numbered list — one variant per line, no explanations, no preamble."
)

_GEN_SYSTEM = (
    "You are a helpful e-commerce data assistant. "
    "Answer questions using only the provided context. "
    "If the answer cannot be found in the context, say so clearly."
)


def _parse_variants(text: str, original: str) -> List[str]:
    variants = []
    for line in text.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        cleaned = re.sub(r"^\d+[.)]\s*|^[-*]\s*", "", line).strip()
        if cleaned:
            variants.append(cleaned)
    seen = {original.strip().lower()}
    unique = [original]
    for v in variants:
        key = v.strip().lower()
        if key and key not in seen:
            seen.add(key)
            unique.append(v)
        if len(unique) >= NUM_QUERY_VARIANTS:
            break
    while len(unique) < NUM_QUERY_VARIANTS:
        unique.append(original)
    return unique[:NUM_QUERY_VARIANTS]


def _expand_query(question: str, primary_idx: int) -> List[str]:
    n_gen = max(1, NUM_QUERY_VARIANTS - 1)
    try:
        resp = _groq_chat(
            primary_idx,
            [
                {"role": "system", "content": _EXPANDER_SYSTEM.format(n=n_gen)},
                {"role": "user",   "content": f"Original question: {question}"},
            ],
            temperature=EXPANDER_TEMPERATURE,
            max_tokens=300,
        )
        return _parse_variants(resp, question)
    except Exception as exc:
        log.warning("Query expansion failed for key#%d: %s", primary_idx + 1, exc)
        return [question] * NUM_QUERY_VARIANTS


def _generate_answer(question: str, docs: list, primary_idx: int) -> str:
    ctx = "\n\n".join(f"[Document {i+1}]\n{d['text']}" for i, d in enumerate(docs))
    return _groq_chat(
        primary_idx,
        [
            {"role": "system", "content": _GEN_SYSTEM},
            {"role": "user",   "content": f"Context:\n{ctx}\n\nQuestion: {question}\n\nAnswer:"},
        ],
        temperature=0.0,
        max_tokens=1024,
    )


# ══════════════════════════════════════════════════════════════════════════════
# Reference-based metric helpers  (zero LLM calls)
# ══════════════════════════════════════════════════════════════════════════════

def _token_list(text: str) -> list:
    return re.sub(r'[^\w\s]', '', (text or "").lower()).split()


def _token_set(text: str) -> set:
    return set(_token_list(text))


def _tfidf_cosine(text_a: str, text_b: str) -> float:
    try:
        m = TfidfVectorizer(min_df=1).fit_transform([text_a or " ", text_b or " "])
        return float(_cosine_sim(m[0], m[1])[0][0])
    except Exception:
        return 0.0


def _split_sentences(text: str) -> list:
    return [s.strip() for s in re.split(r'[.!?]+', (text or "")) if s.strip()]


def _sentence_token_recall(sentence: str, context_tokens: set) -> float:
    s_toks = {t for t in _token_set(sentence) if len(t) > 1}
    if not s_toks:
        return 0.0
    return len(s_toks & context_tokens) / len(s_toks)


def _sentence_faithfulness(generated_answer: str, combined_context: str,
                            support_threshold: float = 0.5) -> float:
    sentences  = _split_sentences(generated_answer)
    if not sentences:
        return 0.0
    ctx_tokens = _token_set(combined_context)
    supported  = sum(
        1 for s in sentences
        if _sentence_token_recall(s, ctx_tokens) >= support_threshold
    )
    return supported / len(sentences)


def _lcs_length(a: list, b: list) -> int:
    prev = [0] * (len(b) + 1)
    for tok_a in a:
        curr = [0] * (len(b) + 1)
        for j, tok_b in enumerate(b, 1):
            curr[j] = prev[j - 1] + 1 if tok_a == tok_b else max(curr[j - 1], prev[j])
        prev = curr
    return prev[len(b)]


def _rouge_l(pred: str, ref: str) -> float:
    p_toks = _token_list(pred)
    r_toks = _token_list(ref)
    if not p_toks or not r_toks:
        return 0.0
    lcs = _lcs_length(p_toks, r_toks)
    if lcs == 0:
        return 0.0
    prec = lcs / len(p_toks)
    rec  = lcs / len(r_toks)
    return 2 * prec * rec / (prec + rec)


# ══════════════════════════════════════════════════════════════════════════════
# Core metric computation
# ══════════════════════════════════════════════════════════════════════════════

def _compute_metrics(
    generated_answer: str,
    retrieved_contexts: list,
    question: str,
    expected_answer: str,
    retrieved_context_ids: list,
    expected_source_ids: list,
) -> dict:
    """All 11 metrics computed without any LLM calls."""
    combined_retrieved = " ".join(retrieved_contexts)
    expected_id_set    = set(expected_source_ids)

    rel_flags = [doc_id in expected_id_set for doc_id in retrieved_context_ids]
    num_rel   = sum(rel_flags)

    faithfulness     = _sentence_faithfulness(generated_answer, combined_retrieved)
    answer_relevancy = _tfidf_cosine(generated_answer, question)

    if not retrieved_context_ids or num_rel == 0:
        context_precision = 0.0
    else:
        ap = sum(
            (sum(rel_flags[:k + 1]) / (k + 1)) * rel_flags[k]
            for k in range(len(rel_flags))
        )
        context_precision = ap / num_rel

    exp_tokens = _token_set(str(expected_answer))
    if not exp_tokens:
        context_recall = 0.0
    else:
        ctx_tokens = _token_set(combined_retrieved)
        context_recall = len(exp_tokens & ctx_tokens) / len(exp_tokens)

    factual_correctness = _rouge_l(str(generated_answer), str(expected_answer))

    ctx_relevancy_scores = [
        _tfidf_cosine(ctx, question) for ctx in retrieved_contexts
    ]
    ctx_relevancy = (
        sum(ctx_relevancy_scores) / len(ctx_relevancy_scores)
        if ctx_relevancy_scores else 0.0
    )
    hallucination = 1.0 - faithfulness

    return {
        "faithfulness":                  faithfulness,
        "answer_relevancy":              answer_relevancy,
        "context_precision":             context_precision,
        "context_recall":                context_recall,
        "factual_correctness(mode=f1)":  factual_correctness,
        "AnswerRelevancyMetric":         answer_relevancy,
        "FaithfulnessMetric":           faithfulness,
        "ContextualPrecisionMetric":    context_precision,
        "ContextualRecallMetric":       context_recall,
        "ContextualRelevancyMetric":    ctx_relevancy,
        "HallucinationMetric":          hallucination,
    }


# ══════════════════════════════════════════════════════════════════════════════
# Vector store setup
# ══════════════════════════════════════════════════════════════════════════════

def _ensure_vector_store():
    try:
        get_collection()
        log.info("ChromaDB index found.")
    except Exception:
        log.info("Building ChromaDB index (~2 min)...")
        build_vector_store()


def _parse_json_list(val: Any) -> List[str]:
    if isinstance(val, list):
        return val
    try:
        r = json.loads(val)
        return r if isinstance(r, list) else [str(r)]
    except Exception:
        return [str(val)] if val else []


def _flush(lines: List[str]) -> None:
    with _print_lock:
        print("\n".join(lines))
        sys.stdout.flush()


# ══════════════════════════════════════════════════════════════════════════════
# Single-query evaluator  (runs inside a worker thread)
# ══════════════════════════════════════════════════════════════════════════════

def _evaluate_one(row, global_idx: int, total: int, primary_idx: int) -> tuple:
    """
    Steps 1-6 for one query. Terminal output buffered and flushed atomically.

    Step 1: Groq query expansion  (1 Groq call)
    Step 2: Multi-retrieval       (PER_QUERY_TOP_K docs per variant, no LLM)
    Step 3: RRF fusion            (merge to FINAL_TOP_K, no LLM)
    Step 4: Groq answer generation (1 Groq call)
    Step 5: RAGAS metrics         (reference-based, no LLM)
    Step 6: DeepEval metrics      (reference-based, no LLM)
    """
    qid        = row.question_id
    difficulty = getattr(row, "difficulty", "-")
    qtype      = getattr(row, "question_type", "-")
    kb_layer   = getattr(row, "best_kb_layer", "-")
    question   = row.question
    expected   = row.expected_answer
    exp_ctx    = _parse_json_list(row.expected_context)
    exp_src    = _parse_json_list(getattr(row, "expected_source_ids", "[]"))

    buf: List[str] = [
        "",
        "=" * _W,
        f"  QUERY {global_idx}/{total}  |  {qid}  |  difficulty: {difficulty}"
        f"  |  type: {qtype}  |  Groq key: #{primary_idx + 1}",
        "=" * _W,
        f"  Question : {question}",
        "",
    ]

    # ── STEP 1: Query Expansion ───────────────────────────────────────────────
    n_gen = NUM_QUERY_VARIANTS - 1
    buf.append(f"  +- STEP 1 . Query Expansion  ({n_gen} variants via Groq, 1 LLM call) "
               f"{'-' * max(0, _W - 55)}")
    expanded_queries: List[str] = []
    try:
        expanded_queries = _expand_query(question, primary_idx)
        buf.append(f"  |  Generated {len(expanded_queries)} variants  "
                   f"(original + {n_gen} paraphrases)")
        for i, q in enumerate(expanded_queries, 1):
            tag = " [original]" if i == 1 else ""
            buf.append(f"  |    [{i}] {q[:90]}{tag}")
        buf.append(f"  +{'-' * (_W - 2)}")
    except Exception as exc:
        buf += [f"  |  ERROR: {exc}", f"  +{'-' * (_W - 2)}",
                f"\n  [SKIP]  Query {global_idx}/{total} [{qid}] -- expansion error\n"]
        _flush(buf)
        empty = {
            "question_id": qid, "difficulty": difficulty, "question_type": qtype,
            "best_kb_layer": kb_layer, "question": question,
            "expected_answer": expected,
            "generated_answer": f"EXPANSION ERROR: {exc}",
            "retrieved_contexts": [], "expected_context": exp_ctx,
            "expanded_queries": [],
        }
        return empty, {"question_id": qid}, {"question_id": qid, "question": question}

    # ── STEP 2: Multi-Retrieval ───────────────────────────────────────────────
    buf.append(f"  +- STEP 2 . Multi-Retrieval  "
               f"({NUM_QUERY_VARIANTS} variants x top-{PER_QUERY_TOP_K}) "
               f"{'-' * max(0, _W - 50)}")
    query_results: dict = {}
    try:
        for q in expanded_queries:
            query_results[q] = retrieve_for_query(q, top_n=PER_QUERY_TOP_K)
        total_raw = sum(len(v) for v in query_results.values())
        buf.append(f"  |  Retrieved {total_raw} raw candidates "
                   f"({NUM_QUERY_VARIANTS} queries x {PER_QUERY_TOP_K})")
        for i, (q, docs) in enumerate(query_results.items(), 1):
            buf.append(f"  |    query [{i}]: {len(docs)} docs  -> "
                       f"{q[:60]}{'...' if len(q) > 60 else ''}")
        buf.append(f"  +{'-' * (_W - 2)}")
    except Exception as exc:
        buf += [f"  |  ERROR: {exc}", f"  +{'-' * (_W - 2)}"]

    # ── STEP 3: RRF Fusion ────────────────────────────────────────────────────
    buf.append(f"  +- STEP 3 . RRF Fusion  (k={RRF_K}, top-{FINAL_TOP_K}) "
               f"{'-' * max(0, _W - 42)}")
    fused_docs = []
    try:
        fused_docs = rrf_fuse(list(query_results.values()), k=RRF_K, top_n=FINAL_TOP_K)
        buf.append(f"  |  Fused {sum(len(v) for v in query_results.values())} raw -> "
                   f"{len(fused_docs)} unique docs  (RRF deduplication)")
        for i, d in enumerate(fused_docs, 1):
            rrf  = d.get("rrf_score", 0)
            dist = d.get("distance", 0)
            src  = d.get("metadata", {}).get("source", d.get("id", ""))
            snip = d["text"].replace("\n", " ")[:80]
            buf.append(f"  |    [{i}] rrf={rrf:.4f}  dist={dist:.4f}  src={src}")
            buf.append(f"  |         +- {snip}...")
        buf.append(f"  +{'-' * (_W - 2)}")
    except Exception as exc:
        buf += [f"  |  ERROR in RRF: {exc}", f"  +{'-' * (_W - 2)}"]
        fused_docs = []

    # ── STEP 4: Groq Generation ───────────────────────────────────────────────
    title4 = f"Groq LLM Generation  ({GROQ_MODEL})  [key #{primary_idx + 1}]"
    buf.append(f"  +- STEP 4 . {title4} {'-' * max(0, _W - 12 - len(title4))}")
    answer = ""
    try:
        answer = _generate_answer(question, fused_docs, primary_idx)
        buf.append(f"  |  Generated answer:")
        for part in textwrap.wrap(answer, width=_W - 10):
            buf.append(f"  |      {part}")
        buf.append(f"  +{'-' * (_W - 2)}")
    except Exception as exc:
        buf += [f"  |  ERROR: {exc}", f"  +{'-' * (_W - 2)}"]
        answer = f"GENERATION ERROR: {exc}"

    rag_r = {
        "question_id":           qid,
        "difficulty":            difficulty,
        "question_type":         qtype,
        "best_kb_layer":         kb_layer,
        "question":              question,
        "expected_answer":       expected,
        "generated_answer":      answer,
        "retrieved_contexts":    [d["text"] for d in fused_docs],
        "retrieved_context_ids": [d.get("id", "") for d in fused_docs],
        "expected_context":      exp_ctx,
        "expected_source_ids":   exp_src,
        "expanded_queries":      expanded_queries,
        "query_results":         {q: [d["id"] for d in docs]
                                  for q, docs in query_results.items()},
    }

    # ── STEP 5: RAGAS Metrics ─────────────────────────────────────────────────
    buf.append(f"  +- STEP 5 . RAGAS Metrics  [reference-based, 0 LLM calls] "
               f"{'-' * 20}")
    buf.append(f"  |  Computing RAGAS metrics vs golden dataset for [{qid}]...")
    buf.append(f"  |  Expected source IDs : {exp_src}")
    buf.append(f"  |  Fused doc IDs       : {rag_r['retrieved_context_ids']}")
    buf.append(f"  |  Expected answer     : {str(expected)[:70]}")

    ragas_row = {"question_id": qid}
    m = {}
    try:
        m = _compute_metrics(
            answer,
            rag_r["retrieved_contexts"],
            question,
            str(expected),
            rag_r["retrieved_context_ids"],
            exp_src,
        )
        buf += ["  |",
                f"  |  {'Metric':<38}  {'Score':>8}  Status  Method",
                f"  |  {'-'*38}  {'-'*8}  {'-'*6}  {'-'*30}"]
        for col in RAGAS_COLS:
            v   = m[col]
            st  = "PASS" if v >= 0.5 else "FAIL"
            mth = RAGAS_METHODS[col]
            buf.append(f"  |  {RAGAS_DISPLAY[col]:<38}  {v:>8.4f}  [{st}]  {mth}")
            ragas_row[col] = v
        buf += ["  |",
                f"  |  [OK]  RAGAS evaluation COMPLETED for query {global_idx}/{total} [{qid}]"]
    except Exception as exc:
        buf.append(f"  |  ERROR: {exc}")
        for col in RAGAS_COLS:
            ragas_row[col] = None
    buf.append(f"  +{'-' * (_W - 2)}")

    # ── STEP 6: DeepEval Metrics ──────────────────────────────────────────────
    buf.append(f"  +- STEP 6 . DeepEval Metrics  [reference-based, 0 LLM calls] "
               f"{'-' * 17}")
    buf.append(f"  |  Computing DeepEval metrics vs golden dataset for [{qid}]...")
    de_row = {"question_id": qid, "question": question}
    try:
        buf += ["  |",
                f"  |  {'Metric':<38}  {'Score':>8}  Status  Computation",
                f"  |  {'-'*38}  {'-'*8}  {'-'*6}  {'-'*38}"]
        for col in DE_COLS:
            v      = m.get(col, 0.0)
            st     = "PASS" if v >= 0.5 else "FAIL"
            reason = DE_REASONS[col]
            rsn    = reason[:55] + ("..." if len(reason) > 55 else "")
            buf.append(f"  |  {DE_DISPLAY[col]:<38}  {v:>8.4f}  [{st}]  {rsn}")
            de_row[col]             = v
            de_row[col + "_reason"] = reason
        buf += ["  |",
                f"  |  [OK]  DeepEval evaluation COMPLETED for query {global_idx}/{total} [{qid}]"]
    except Exception as exc:
        buf.append(f"  |  ERROR: {exc}")
        for col in DE_COLS:
            de_row[col]             = None
            de_row[col + "_reason"] = str(exc)

    pipeline_desc = (f"Expand({NUM_QUERY_VARIANTS}) -> "
                     f"MultiRetrieve({PER_QUERY_TOP_K}ea) -> "
                     f"RRF({FINAL_TOP_K}) -> Generate -> RAGAS -> DeepEval")
    buf += [
        f"  +{'-' * (_W - 2)}",
        "",
        f"  [END-TO-END COMPLETE]  Query {global_idx}/{total} [{qid}]"
        f"  |  key #{primary_idx + 1}  |  {pipeline_desc}",
        "",
    ]

    _flush(buf)
    return rag_r, ragas_row, de_row


# ══════════════════════════════════════════════════════════════════════════════
# Batch runner
# ══════════════════════════════════════════════════════════════════════════════

def _run_batch(batch_rows: list, global_offsets: List[int], total: int,
               primary_idx: int) -> List[tuple]:
    results: List[tuple] = []
    n = len(batch_rows)
    for local_i, (row, g_idx) in enumerate(zip(batch_rows, global_offsets)):
        rag_r, ragas_r, de_r = _evaluate_one(row, g_idx, total, primary_idx)
        results.append((g_idx, rag_r, ragas_r, de_r))
        if local_i < n - 1:
            # Longer delay: 2 Groq calls per query vs. 1 for other pipelines
            delay = random.uniform(10, 25)
            with _print_lock:
                print(f"  [sleep]  [key #{primary_idx + 1}] Sleeping {delay:.1f}s "
                      f"before next query...")
                sys.stdout.flush()
            time.sleep(delay)
    return results


# ══════════════════════════════════════════════════════════════════════════════
# Parallel evaluation coordinator
# ══════════════════════════════════════════════════════════════════════════════

def evaluate_all(golden_df: pd.DataFrame, batch_size: int = 20,
                 sparse_n: int = None) -> tuple:
    total_dataset = len(golden_df)
    eval_mode     = "reference-based (token-overlap + TF-IDF, no LLM judge)"
    rows_list     = list(golden_df.itertuples(index=False))

    batches: List[tuple] = []
    if sparse_n:
        for k in range(_PRIMARY_KEY_COUNT):
            start  = k * batch_size
            b_rows = []
            b_off  = []
            for i in range(sparse_n):
                idx = start + i
                if idx < total_dataset:
                    b_rows.append(rows_list[idx])
                    b_off.append(idx + 1)
            if b_rows:
                batches.append((b_rows, b_off, k))
    else:
        for g_idx, row in enumerate(rows_list):
            p_idx = _primary_key_idx(g_idx, batch_size)
            while len(batches) <= p_idx:
                batches.append(([], [], len(batches)))
            batches[p_idx][0].append(row)
            batches[p_idx][1].append(g_idx + 1)

    n_selected = sum(len(b[0]) for b in batches)

    print(f"\n{'#' * _W}")
    print(f"#  END-TO-END MULTI-QUERY RAG EVALUATION  --  "
          f"{n_selected} queries selected  (dataset size: {total_dataset})")
    print(f"#  Pipeline         : Expand({NUM_QUERY_VARIANTS}) -> "
          f"MultiRetrieve({PER_QUERY_TOP_K}ea) -> RRF({FINAL_TOP_K}) -> Groq")
    print(f"#  LLM (expansion)  : {GROQ_MODEL}  (temp={EXPANDER_TEMPERATURE})")
    print(f"#  LLM (generation) : {GROQ_MODEL}  (temp=0.0)")
    print(f"#  Groq calls/query : 2  (expansion + generation)")
    print(f"#  Evaluation method: {eval_mode}")
    if sparse_n:
        print(f"#  Mode             : SPARSE  --  {sparse_n} queries per key")
    else:
        print(f"#  Key layout       : 1-{batch_size} -> key#1 | "
              f"{batch_size+1}-{2*batch_size} -> key#2 | ...")
    print(f"#  Fallbacks        : keys #6-#{len(GROQ_API_KEYS)} auto-used when primary exhausted")
    print(f"#  Delay            : random 10-25 s between queries (2 LLM calls each)")
    print(f"{'#' * _W}")

    print(f"\n  Batches: {len(batches)}  |  Parallel workers: {len(batches)}")
    for b_rows, b_off, p_idx in batches:
        positions = ", ".join(str(x) for x in b_off)
        print(f"    Groq key #{p_idx + 1}  ->  dataset positions [{positions}]  ({len(b_rows)} queries)")
    print()

    all_raw: List[tuple] = []
    with ThreadPoolExecutor(max_workers=len(batches)) as executor:
        futures = {
            executor.submit(_run_batch, b_rows, b_off, total_dataset, p_idx): p_idx
            for (b_rows, b_off, p_idx) in batches
        }
        for fut in as_completed(futures):
            p_idx = futures[fut]
            try:
                batch_res = fut.result()
                all_raw.extend(batch_res)
                with _print_lock:
                    print(f"\n  [OK]  Batch key#{p_idx + 1} finished -- "
                          f"{len(batch_res)} queries done")
                    sys.stdout.flush()
            except Exception as exc:
                with _print_lock:
                    log.error("Batch key#%d raised: %s", p_idx + 1, exc)

    all_raw.sort(key=lambda t: t[0])
    rag_rows   = [t[1] for t in all_raw]
    ragas_rows = [t[2] for t in all_raw]
    de_rows    = [t[3] for t in all_raw]

    return rag_rows, ragas_rows, de_rows, eval_mode


# ══════════════════════════════════════════════════════════════════════════════
# Excel export
# ══════════════════════════════════════════════════════════════════════════════

def _save_excel(rag_rows, ragas_rows, de_rows, eval_mode, out_path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    BRAND      = "7B1FA2"    # purple — multi-query brand colour
    BRAND_LITE = "F3E5F5"
    DARK       = "212121"
    PASS_GREEN = "C8E6C9"
    FAIL_RED   = "FFCDD2"
    GREY_HDR   = "37474F"
    TOTAL_GREY = "ECEFF1"

    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hf(hex_c):
        return PatternFill("solid", fgColor=hex_c)

    def _cell(ws, r, c, val="", bold=False, color=None, bg=None,
              halign="center", wrap=False):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font      = Font(bold=bold, color=color or "000000", size=10)
        cell.alignment = Alignment(horizontal=halign, vertical="center",
                                   wrap_text=wrap)
        cell.border    = bdr
        if bg:
            cell.fill = _hf(bg)
        return cell

    ragas_df = pd.DataFrame(ragas_rows).set_index("question_id")
    de_df    = pd.DataFrame(de_rows).set_index("question_id")
    rag_df   = pd.DataFrame(rag_rows).set_index("question_id")

    ragas_means = ragas_df[[c for c in RAGAS_COLS if c in ragas_df]].mean()
    de_means    = de_df[[c for c in DE_COLS if c in de_df]].mean()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Combined Results ─────────────────────────────────────────────
    ws1 = wb.create_sheet("Combined Results")
    metric_cols  = [RAGAS_DISPLAY[c] for c in RAGAS_COLS] + [DE_DISPLAY[c] for c in DE_COLS]
    all_cols     = ["question_id", "question"] + metric_cols
    ws1.append(all_cols)
    for cell in ws1[1]:
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = _hf(BRAND)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr
    ws1.row_dimensions[1].height = 20
    for row in rag_df.itertuples():
        qid = row.Index
        q   = getattr(row, "question", "")
        vals = [qid, q]
        for rc in RAGAS_COLS:
            v = ragas_df.at[qid, rc] if qid in ragas_df.index and rc in ragas_df.columns else ""
            vals.append(round(v, 4) if isinstance(v, float) else v)
        for dc in DE_COLS:
            v = de_df.at[qid, dc] if qid in de_df.index and dc in de_df.columns else ""
            vals.append(round(v, 4) if isinstance(v, float) else v)
        ws1.append(vals)
    # FINAL row
    final_vals = ["FINAL", "FINAL (mean)"]
    for rc in RAGAS_COLS:
        final_vals.append(round(ragas_means.get(rc, 0.0), 3))
    for dc in DE_COLS:
        final_vals.append(round(de_means.get(dc, 0.0), 3))
    ws1.append(final_vals)
    fr = ws1.max_row
    for c_idx in range(1, len(all_cols) + 1):
        cell = ws1.cell(row=fr, column=c_idx)
        cell.font   = Font(bold=True, size=10)
        cell.fill   = _hf(TOTAL_GREY)
        cell.border = bdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.freeze_panes = "A2"
    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 55

    # ── Sheet 2: RAG Responses ────────────────────────────────────────────────
    ws2 = wb.create_sheet("RAG Responses")
    resp_hdr = ["question_id", "question", "expanded_queries",
                "generated_answer", "expected_answer",
                "retrieved_context_ids", "retrieved_contexts"]
    ws2.append(resp_hdr)
    for cell in ws2[1]:
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = _hf(BRAND)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = bdr
    for row in rag_df.itertuples():
        eq   = getattr(row, "expanded_queries", [])
        eq_s = " | ".join(eq) if isinstance(eq, list) else str(eq)
        ctx  = getattr(row, "retrieved_contexts", [])
        ctx_s = "\n---\n".join(ctx) if isinstance(ctx, list) else str(ctx)
        cids = getattr(row, "retrieved_context_ids", [])
        cids_s = ", ".join(cids) if isinstance(cids, list) else str(cids)
        ws2.append([
            row.Index,
            getattr(row, "question", ""),
            eq_s,
            getattr(row, "generated_answer", ""),
            getattr(row, "expected_answer", ""),
            cids_s,
            ctx_s,
        ])
    ws2.freeze_panes = "A2"
    for i, w in enumerate([14, 55, 55, 55, 40, 40, 80], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: RAGAS Metrics ────────────────────────────────────────────────
    ws3 = wb.create_sheet("RAGAS Metrics")
    ragas_hdr = ["question_id"] + [RAGAS_DISPLAY[c] for c in RAGAS_COLS]
    ws3.append(ragas_hdr)
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = _hf(BRAND)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr
    for r_row in ragas_rows:
        qid = r_row.get("question_id", "")
        vals = [qid]
        for col in RAGAS_COLS:
            v = r_row.get(col, "")
            vals.append(round(v, 4) if isinstance(v, float) else v)
        ws3.append(vals)
        data_row = ws3.max_row
        for ci, col in enumerate(RAGAS_COLS, 2):
            v = r_row.get(col)
            if isinstance(v, float):
                ws3.cell(data_row, ci).fill = _hf(PASS_GREEN if v >= 0.5 else FAIL_RED)
    ws3.freeze_panes = "A2"

    # ── Sheet 4: DeepEval Metrics ─────────────────────────────────────────────
    ws4 = wb.create_sheet("DeepEval Metrics")
    de_hdr = ["question_id", "question"] + [DE_DISPLAY[c] for c in DE_COLS]
    ws4.append(de_hdr)
    for cell in ws4[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = _hf(BRAND)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr
    for d_row in de_rows:
        qid = d_row.get("question_id", "")
        vals = [qid, d_row.get("question", "")]
        for col in DE_COLS:
            v = d_row.get(col, "")
            vals.append(round(v, 4) if isinstance(v, float) else v)
        ws4.append(vals)
        data_row = ws4.max_row
        for ci, col in enumerate(DE_COLS, 3):
            v = d_row.get(col)
            if isinstance(v, float):
                ws4.cell(data_row, ci).fill = _hf(PASS_GREEN if v >= 0.5 else FAIL_RED)
    ws4.freeze_panes = "A2"

    # ── Sheet 5: Summary ──────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Summary")
    ws5.merge_cells("A1:E1")
    ws5["A1"].value = f"{RAG_NAME} — Evaluation Summary"
    ws5["A1"].font  = Font(bold=True, size=13, color="FFFFFF")
    ws5["A1"].fill  = _hf(DARK)
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 26

    ws5.merge_cells("A2:E2")
    ws5["A2"].value = (f"Pipeline: Expand({NUM_QUERY_VARIANTS}) -> "
                       f"MultiRetrieve({PER_QUERY_TOP_K}ea) -> RRF({FINAL_TOP_K}) -> "
                       f"{GROQ_MODEL}")
    ws5["A2"].font  = Font(italic=True, size=9, color="555555")
    ws5["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws5.merge_cells("A3:E3")
    ws5["A3"].value = f"Evaluation: {eval_mode}"
    ws5["A3"].font  = Font(italic=True, size=9, color="555555")
    ws5["A3"].alignment = Alignment(horizontal="center", vertical="center")

    # META rows
    meta_r = 4
    for key, val in [
        ("RAG Type",          RAG_NAME),
        ("LLM Model",         GROQ_MODEL),
        ("Query Variants",    NUM_QUERY_VARIANTS),
        ("Per-Variant Top-k", PER_QUERY_TOP_K),
        ("Final Top-k (RRF)", FINAL_TOP_K),
        ("RRF k constant",    RRF_K),
        ("Expander Temp",     EXPANDER_TEMPERATURE),
        ("Groq calls/query",  2),
        ("Eval Method",       eval_mode),
    ]:
        ws5.cell(meta_r, 1, key).font = Font(bold=True, size=10)
        ws5.cell(meta_r, 1).fill = _hf(BRAND_LITE)
        ws5.cell(meta_r, 1).alignment = Alignment(horizontal="left", vertical="center")
        ws5.cell(meta_r, 1).border = bdr
        ws5.merge_cells(start_row=meta_r, start_column=2, end_row=meta_r, end_column=5)
        ws5.cell(meta_r, 2, str(val)).alignment = Alignment(horizontal="left")
        ws5.cell(meta_r, 2).border = bdr
        meta_r += 1

    meta_r += 1  # blank separator
    hdr_r = meta_r
    for ci, hdr in enumerate(["Metric", "Mean", "Min", "Max", "Std"], 1):
        c = ws5.cell(hdr_r, ci, hdr)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = _hf(GREY_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr
    ws5.row_dimensions[hdr_r].height = 20

    stat_r = hdr_r + 1
    all_metric_cols = (
        [(RAGAS_DISPLAY[c], ragas_df[c]) for c in RAGAS_COLS if c in ragas_df.columns]
        + [(DE_DISPLAY[c],  de_df[c])    for c in DE_COLS    if c in de_df.columns]
    )
    for display_name, series in all_metric_cols:
        numeric = pd.to_numeric(series, errors="coerce").dropna()
        mn   = round(numeric.mean(), 4)  if len(numeric) else ""
        mi   = round(numeric.min(),  4)  if len(numeric) else ""
        mx   = round(numeric.max(),  4)  if len(numeric) else ""
        std  = round(numeric.std(),  4)  if len(numeric) else ""
        ws5.cell(stat_r, 1, display_name).font   = Font(size=10)
        ws5.cell(stat_r, 1).border                = bdr
        ws5.cell(stat_r, 1).alignment             = Alignment(horizontal="left")
        ws5.cell(stat_r, 1).fill                  = _hf(BRAND_LITE)
        for ci, v in enumerate([mn, mi, mx, std], 2):
            cell = ws5.cell(stat_r, ci, v)
            cell.border    = bdr
            cell.alignment = Alignment(horizontal="center")
            if isinstance(v, float):
                cell.fill = _hf(PASS_GREEN if v >= 0.5 else FAIL_RED)
        stat_r += 1

    for i, w in enumerate([35, 12, 12, 12, 12], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    log.info("Saved -> %s", out_path)


# ══════════════════════════════════════════════════════════════════════════════
# Aggregate score printer
# ══════════════════════════════════════════════════════════════════════════════

def _print_aggregates(ragas_rows, de_rows) -> None:
    ragas_df = pd.DataFrame(ragas_rows)
    de_df    = pd.DataFrame(de_rows)

    print(f"\n{'#' * _W}")
    print(f"#  ALL QUERIES COMPLETE -- AGGREGATE SCORES")
    print(f"{'#' * _W}")
    print()
    print("  RAGAS (mean across all queries):")
    for col in RAGAS_COLS:
        if col in ragas_df.columns:
            v = pd.to_numeric(ragas_df[col], errors="coerce").mean()
            print(f"    {RAGAS_DISPLAY[col]:<40}  {v:.4f}")
    print()
    print("  DeepEval (mean across all queries):")
    for col in DE_COLS:
        if col in de_df.columns:
            v = pd.to_numeric(de_df[col], errors="coerce").mean()
            print(f"    {DE_DISPLAY[col]:<40}  {v:.4f}")
    print()


# ══════════════════════════════════════════════════════════════════════════════
# Entry point
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    parser = argparse.ArgumentParser(description="Multi-Query RAG end-to-end evaluation")
    parser.add_argument("--limit",      type=int, default=None,
                        help="Evaluate only first N rows of the golden dataset")
    parser.add_argument("--batch-size", type=int, default=20,
                        help="Queries per Groq-key batch (default 20)")
    parser.add_argument("--sparse",     type=int, default=None,
                        help="Run only first N queries per key (sparse mode)")
    args = parser.parse_args()

    out_path = _default_output()
    log.info("Output file: %s", out_path.name)

    golden_df = pd.read_csv(GOLDEN_PATH)
    if args.limit:
        golden_df = golden_df.head(args.limit)
    log.info("Dataset rows loaded: %d", len(golden_df))

    _ensure_vector_store()

    rag_rows, ragas_rows, de_rows, eval_mode = evaluate_all(
        golden_df,
        batch_size=args.batch_size,
        sparse_n=args.sparse,
    )

    log.info("Writing Excel -> %s", out_path)
    _save_excel(rag_rows, ragas_rows, de_rows, eval_mode, out_path)

    _print_aggregates(ragas_rows, de_rows)

    ragas_df = pd.DataFrame(ragas_rows)
    de_df    = pd.DataFrame(de_rows)
    all_data: dict = {"question_id": [r["question_id"] for r in rag_rows]}
    for r in rag_rows:
        all_data.setdefault("question", []).append(r.get("question", ""))
    for col in RAGAS_COLS:
        if col in ragas_df.columns:
            all_data[RAGAS_DISPLAY[col]] = list(ragas_df[col])
    for col in DE_COLS:
        if col in de_df.columns:
            all_data[DE_DISPLAY[col]] = list(de_df[col])
    result_df = pd.DataFrame(all_data).set_index("question_id")
    mean_row  = result_df.select_dtypes("number").mean().rename("FINAL")
    result_df = pd.concat([result_df, mean_row.to_frame().T])
    result_df.index.name = "question_id"

    print()
    print("=" * (_W + 10))
    print("EVALUATION RESULTS -- each row = one query, each column = one metric")
    print("=" * (_W + 10))
    print(result_df.to_string())
    print(f"\n  Eval method : {eval_mode}")
    print(f"  Results     : {out_path}")
    print()


if __name__ == "__main__":
    main()
